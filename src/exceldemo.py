#!/usr/bin/python
# -*- coding:utf-8 -*-
# Author:wydpp

'''
openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
pip install openpyxl
'''

from openpyxl import load_workbook
from openpyxl import Workbook

def readExcelXlsx():
    #打开excel文档
    wb = load_workbook("/home/dpp/Desktop/code/python-demo/resources/exceldemo.xlsx")
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_col=1, max_col=2, min_row=1, max_row=3):
        for c in row:
            #查看对象结构
            #print(dir(c))
            print("column=%d,row=%d,value=%s"%(c.column,c.row,c.value))

def writeExcel():
    wb = Workbook()
    ws = wb.create_sheet("pythonSheet")
    for row in ws.iter_rows(min_col=1, max_col=2, min_row=1, max_row=3):
        for c in row:
            #给单元格赋值
            c.value= str(c.row)+"行，"+str(c.column)+"列"
    wb.save("/home/dpp/Desktop/code/python-demo/resources/pythonWriteDemo.xlsx")

writeExcel()